import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
import sys
import os
import datetime
import json



def send_filelist(msg_content, file_paths):
    subject = '实验结果' + (datetime.datetime.now().strftime('%Y-%m-%d:%H:%M'))
    # 发信方的信息：发信邮箱，QQ 邮箱授权码
    from_addr = 'tangweize'
    password = '********1******'
    # 收信方邮箱
    recievers = ['756648174@qq.com']
    # 发信服务器
    smtp_server = 'smtp.exmail.qq.com'

    # 创建邮件
    msg = MIMEMultipart()
    msg['From'] = Header('读数完成')  # 发送者
    msg['Subject'] = Header(subject, 'utf-8')  # 邮件主题
    msg.attach(MIMEText(msg_content, 'plain', 'utf-8'))

    # **遍历所有文件并添加到邮件附件**
    for file_path in file_paths:
        try:
            with open(file_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment', filename=file_path.split('/')[-1])
                msg.attach(part)
        except Exception as e:
            print(f"无法附加文件 {file_path}: {e}")

    # 发送邮件
    try:
        smtpobj = smtplib.SMTP_SSL(smtp_server)
        smtpobj.connect(smtp_server, 465)
        smtpobj.login(from_addr, password)
        smtpobj.sendmail(from_addr, recievers, msg.as_string())
        print("邮件发送成功")
    except smtplib.SMTPException as e:
        print(f"无法发送邮件: {e}")
    finally:
        smtpobj.quit()

def get_delta_date_str(date_str, delta):
    return  (datetime.datetime.strptime(date_str, "%Y-%m-%d") + datetime.timedelta(days=delta)).strftime('%Y-%m-%d')

def str2date(date_str):
    return datetime.datetime.strptime(date_str, "%Y-%m-%d")

def date2str(date):
    return date.strftime('%Y-%m-%d')

def get_train_test_data(data, startdate, enddate, testdate):
    train_data = data[
        (data['install_date'] >= startdate) & (data['install_date'] <= enddate)].reset_index(
        drop=True).copy() 

    test_data = data[
        (data['install_date'] == testdate)].reset_index(drop=True).copy()  

    return train_data, test_data

def calculate_bias(data, suffix = 'base'):
    
    channel_site_bias = data.groupby(['install_date', 'channel', 'site_name']).agg({'tweedie_predict_delta_7d':'sum', 'b2_sale_amt_7d':'sum'}).reset_index()
    dim_os_creative_bias = data.groupby(['install_date', 'channel',  'dim_os_name', 'creative_classify']).agg({'tweedie_predict_delta_7d':'sum', 'b2_sale_amt_7d':'sum'}).reset_index()
    channel_site_bias['tweedie_predict_delta_7d' + suffix] =  channel_site_bias['tweedie_predict_delta_7d']
    dim_os_creative_bias['tweedie_predict_delta_7d' + suffix] = dim_os_creative_bias['tweedie_predict_delta_7d']
    
    return channel_site_bias, dim_os_creative_bias 


def merge_results(base_res, iter_res):
    online_channel_site_bias, online_dim_os_creative_bias = calculate_bias(base_res, '_base')
    optim_channel_site_bias, optim_dim_os_creative_bias = calculate_bias(iter_res, '_iter')
    
    channel_site_bias_compare = online_channel_site_bias.merge(optim_channel_site_bias[['install_date', 'channel', 'site_name','tweedie_predict_delta_7d_iter']], 
                                                           on = ['install_date', 'channel', 'site_name', ], 
                                                           how = 'left')
    dim_os_creative_bias = online_dim_os_creative_bias.merge(optim_dim_os_creative_bias[['install_date',  'channel', 'dim_os_name', 'creative_classify', 'tweedie_predict_delta_7d_iter']], 
                                                             on = ['install_date', 'channel', 'dim_os_name', 'creative_classify' ],
                                                             how = 'left')


    channel_site_bias_compare['iter_bias'] = channel_site_bias_compare.apply(lambda row: (row['tweedie_predict_delta_7d_iter'] - row['b2_sale_amt_7d']) / ( row['b2_sale_amt_7d'] + 1), axis = 1 )
    channel_site_bias_compare['base_bias'] = channel_site_bias_compare.apply(lambda row: (row['tweedie_predict_delta_7d_base'] - row['b2_sale_amt_7d']) / ( row['b2_sale_amt_7d'] + 1), axis = 1 )

    dim_os_creative_bias['iter_bias'] = dim_os_creative_bias.apply(lambda row: round((row['tweedie_predict_delta_7d_iter'] - row['b2_sale_amt_7d']) / ( row['b2_sale_amt_7d'] + 1), 3), axis = 1 )
    dim_os_creative_bias['base_bias'] = dim_os_creative_bias.apply(lambda row: round( (row['tweedie_predict_delta_7d_base'] - row['b2_sale_amt_7d']) / ( row['b2_sale_amt_7d'] + 1), 3), axis = 1 )
    
    return channel_site_bias_compare, dim_os_creative_bias

def get_channel_bias(res, thd = -1, base_col = 'tweedie_predict_delta_7d_base', iter_col = 'tweedie_predict_delta_7d_iter', channel = ['今日头条2.0']):
    res = res[res.channel.isin(channel)]
    
    if thd >= 0:
        res = res[res.b2_sale_amt_7d >= thd]
    res = res[(res.base_bias.isna() == False) & (res.iter_bias.isna() == False)]

    res['iter_bias'] = res['iter_bias'].apply(lambda x: abs(x))
    res['base_bias'] = res['base_bias'].apply(lambda x: abs(x))
    
    
    temp = res.groupby(['install_date','channel']).agg({'b2_sale_amt_7d':'sum'})
    res = res.merge(temp.rename(columns={'b2_sale_amt_7d': 'b2_sale_amt_7d_sum'}), 
    on=['install_date','channel'], 
    how='left'
    )
    
    res['sale_weight'] = res['b2_sale_amt_7d'] / res['b2_sale_amt_7d_sum'] 
    
    by_dt_bias = res.groupby('install_date').agg({'base_bias': 'mean', 'iter_bias':'mean'})
    # print("="* 10, "平均bias","="* 10)
    # display(by_dt_bias)
    # display(by_dt_bias.agg({'base_bias': 'mean', 'iter_bias':'mean'}))
    
    
    res['sale_weight'] = res['b2_sale_amt_7d'] / res['b2_sale_amt_7d_sum'] 
    res['base_bias'] = res['base_bias'].apply(lambda x: abs(x)) *  res['sale_weight']
    res['iter_bias'] = res['iter_bias'].apply(lambda x: abs(x)) *  res['sale_weight']


    by_dt_bias = res.groupby('install_date').agg({'base_bias': 'sum', 'iter_bias':'sum'})
    
    by_dt_bias['提升'] = (by_dt_bias['iter_bias'] - by_dt_bias['base_bias']) / (by_dt_bias['base_bias'] + 0.0001)
    
    value_row = pd.DataFrame({
    'base_bias': [by_dt_bias['base_bias'].mean()],
    'iter_bias': [by_dt_bias['iter_bias'].mean()],
    '提升': [(by_dt_bias['iter_bias'].mean() - by_dt_bias['base_bias'].mean()) /( by_dt_bias['base_bias'].mean() + 0.0001) ],
    }, index=['MEAN'])
    
    by_dt_bias = pd.concat([by_dt_bias, value_row])
    
    print("="* 10, "加权bias","="* 10)
    display(by_dt_bias)
    display(by_dt_bias.agg({'base_bias': 'mean', 'iter_bias':'mean'}))
    return by_dt_bias


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import DataBarRule

def save_excel_with_progress_bar(df, column, filename="highlighted.xlsx"):
    """
    在 Excel 里高亮指定 `column`，并为 `base_bias` 和 `iter_bias` 添加进度条，并将其格式化为百分比。
    对“提升”列的正数标柔和红色，负数标柔和绿色。

    参数:
    - df: pandas DataFrame
    - column: 需要高亮的列
    - filename: 输出 Excel 文件名 (默认: "highlighted.xlsx")
    """
    wb = Workbook()
    ws = wb.active

    # 颜色映射
    colors = ["FFDDC1", "FFABAB", "FFC3A0", "D5AAFF", "85E3FF", "B9FBC0"]
    unique_values = df[column].unique()
    color_map = {val: colors[i % len(colors)] for i, val in enumerate(unique_values)}

    # 写入表头
    for col_num, column_title in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_num, value=column_title)

    # 写入数据并高亮
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        value = getattr(row, column)
        fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")

        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.fill = fill

            # 格式化 base_bias 和 iter_bias 为百分比
            if df.columns[col_idx-1] in ["base_bias", "iter_bias","提升"]:
                cell.number_format = '0.00%'  # 将数字格式设置为百分比，保留两位小数

            # 对 “提升” 列进行颜色标记
            if df.columns[col_idx-1] == "提升":
                if cell_value > 0:
                    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # 正数标柔和红色
                elif cell_value < 0:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 负数标柔和绿色

    # 进度条：base_bias 和 iter_bias
    base_bias_col = df.columns.get_loc("base_bias") + 1
    iter_bias_col = df.columns.get_loc("iter_bias") + 1

    max_base = 1.0
    max_iter = 1.0

    base_bias_rule = DataBarRule(start_type="num", start_value=0, 
                                 end_type="num", end_value=max_base, color="63C384")
    iter_bias_rule = DataBarRule(start_type="num", start_value=0, 
                                 end_type="num", end_value=max_iter, color="FF6961")

    ws.conditional_formatting.add(f"{chr(64+base_bias_col)}2:{chr(64+base_bias_col)}{len(df)+1}", base_bias_rule)
    ws.conditional_formatting.add(f"{chr(64+iter_bias_col)}2:{chr(64+iter_bias_col)}{len(df)+1}", iter_bias_rule)

    # 保存 Excel
    wb.save(filename)
    print(f"Excel 文件已保存: {filename}")

# 示例调用
# save_excel_with_progress_bar(df, 'channel', 'highlighted_with_progress.xlsx')


def get_channel_site_bias(res, thd = -1, base_col = 'tweedie_predict_delta_7d_base', iter_col = 'tweedie_predict_delta_7d_iter', channel = ['今日头条2.0']):
    res = res[res.channel.isin(channel)]
    
    if thd >= 0:
        res = res[res.b2_sale_amt_7d >= thd]
    res = res[(res.base_bias.isna() == False) & (res.iter_bias.isna() == False)]

    res['iter_bias'] = res['iter_bias'].apply(lambda x: abs(x))
    res['base_bias'] = res['base_bias'].apply(lambda x: abs(x))
    
    temp = res.groupby(['site_name']).agg({'b2_sale_amt_7d':'sum'})
    res = res.merge(temp.rename(columns={'b2_sale_amt_7d': 'b2_sale_amt_7d_sum'}), 
    on=['site_name'], 
    how='left'
    )

    
    res['提升'] = (res['iter_bias'] - res['base_bias']) / (res['base_bias'] + 0.0001)

    res = res[res.b2_sale_amt_7d_sum > 10000]
    res = res.sort_values(by = ['b2_sale_amt_7d_sum', 'site_name', 'install_date'], ascending = False)
    by_dt_res = res
    summary_res = res.groupby(['channel', 'site_name','b2_sale_amt_7d_sum']).agg({'base_bias': 'mean', 'iter_bias':'mean'}).sort_values(by = ['b2_sale_amt_7d_sum',], ascending = False)

    return by_dt_res, summary_res.reset_index()



def get_channel_biz_os_bias(res, thd = -1, base_col = 'tweedie_predict_delta_7d_base', iter_col = 'tweedie_predict_delta_7d_iter', channel = ['今日头条2.0']):
    res = res[res.channel.isin(channel)]
    
    if thd >= 0:
        res = res[res.b2_sale_amt_7d >= thd]
    res = res[(res.base_bias.isna() == False) & (res.iter_bias.isna() == False)]

    res['iter_bias'] = res['iter_bias'].apply(lambda x: abs(x))
    res['base_bias'] = res['base_bias'].apply(lambda x: abs(x))
    
    temp = res.groupby(['dim_os_name','creative_classify']).agg({'b2_sale_amt_7d':'sum'})
    res = res.merge(temp.rename(columns={'b2_sale_amt_7d': 'b2_sale_amt_7d_sum'}), 
    on=['dim_os_name','creative_classify'], 
    how='left'
    )

    
    res['提升'] = (res['iter_bias'] - res['base_bias']) / (res['base_bias'] + 0.0001)

    res = res[res.b2_sale_amt_7d_sum > 10000]
    res = res.sort_values(by = ['b2_sale_amt_7d_sum', 'creative_classify', 'dim_os_name', 'install_date'], ascending = False)
    by_dt_res = res
    summary_res = res.groupby(['b2_sale_amt_7d_sum','dim_os_name','creative_classify']).agg({'base_bias': 'mean', 'iter_bias':'mean'}).sort_values(by = ['b2_sale_amt_7d_sum',], ascending = False)

    return by_dt_res, summary_res.reset_index()

    
def get_channel_site_bias(res, thd = -1, base_col = 'tweedie_predict_delta_7d_base', iter_col = 'tweedie_predict_delta_7d_iter', channel = ['今日头条2.0']):
    res = res[res.channel.isin(channel)]
    
    if thd >= 0:
        res = res[res.b2_sale_amt_7d >= thd]
    res = res[(res.base_bias.isna() == False) & (res.iter_bias.isna() == False)]

    res['iter_bias'] = res['iter_bias'].apply(lambda x: abs(x))
    res['base_bias'] = res['base_bias'].apply(lambda x: abs(x))
    
    temp = res.groupby(['site_name']).agg({'b2_sale_amt_7d':'sum'})
    res = res.merge(temp.rename(columns={'b2_sale_amt_7d': 'b2_sale_amt_7d_sum'}), 
    on=['site_name'], 
    how='left'
    )

    
    res['提升'] = (res['iter_bias'] - res['base_bias']) / (res['base_bias'] + 0.0001)

    res = res[res.b2_sale_amt_7d_sum > 10000]
    res = res.sort_values(by = ['b2_sale_amt_7d_sum', 'site_name', 'install_date'], ascending = False)
    by_dt_res = res
    summary_res = res.groupby(['channel', 'site_name','b2_sale_amt_7d_sum']).agg({'base_bias': 'mean', 'iter_bias':'mean'}).sort_values(by = ['b2_sale_amt_7d_sum',], ascending = False)

    return by_dt_res, summary_res.reset_index()

import pickle

def save_obj(obj, file_path):
    with open(file_path, 'wb') as f:
        pickle.dump(obj, f)


import pickle

def load_obj(file_path):
    """
    使用 pickle 加载保存的 Python 对象（例如 .pkl 或 .obj 文件）

    参数:
        file_path (str): 文件路径，通常是 .pkl 或 .obj 后缀

    返回:
        object: 从文件中反序列化得到的 Python 对象
    """
    with open(file_path, 'rb') as f:
        obj = pickle.load(f)
    return obj




def read_feature_json_config(filename):
    with open(filename, 'r', encoding='utf-8') as file:
        # 使用json.load()函数读取文件内容并将其转换为字典
        config = json.load(file)
    group_2_features = {}
    for i in config:
        group_2_features[i] = []
        for feature in config[i]:
            group_2_features[i].append(feature['name'])
    return group_2_features


import tensorflow as tf
def tfrecords_to_dataset(path, parse_function, batch_size):
    dataset = tf.data.TFRecordDataset(path)
    dataset = dataset.map(parse_function)

    dataset = dataset.prefetch(buffer_size=10000)
    dataset = dataset.batch(batch_size)

    return dataset



def get_trian_valid_test_dateset(parse_function, batch_size, train_path, valid_path = None, test_path = None ):
    train_dataset = tfrecords_to_dataset(train_path, parse_function, batch_size)
    valid_dataset = None
    test_dataset = None
    if valid_path:
        valid_dataset = tfrecords_to_dataset(valid_path, parse_function, batch_size)
    if test_path:
        test_dataset = tfrecords_to_dataset(test_path, parse_function, batch_size)

    return train_dataset, valid_dataset, test_dataset



import matplotlib.pyplot as plt
import numpy as np

def calculate_area_under_gain_curve(pred_list, true_list, head_name=""):
    # 将零维张量列表转换为一维 NumPy 数组
    pred = pred_list.numpy()
    true = true_list.numpy()

    # 创建 DataFrame
    df = pd.DataFrame({'pred': pred, 'true': true})

    # 【1】预测值排序的增益曲线
    df_pred_sorted = df.sort_values(by='pred', ascending=False).copy()
    df_pred_sorted['cumulative_percentage_customers'] = np.arange(1, len(df_pred_sorted) + 1) / len(df_pred_sorted)
    df_pred_sorted['cumulative_percentage_ltv'] = df_pred_sorted['true'].cumsum() / df_pred_sorted['true'].sum()
    area_pred = np.trapz(df_pred_sorted['cumulative_percentage_ltv'],
                         df_pred_sorted['cumulative_percentage_customers'])

    # 【2】真实值排序的理想增益曲线（Ground Truth 理想线）
    df_true_sorted = df.sort_values(by='true', ascending=False).copy()
    df_true_sorted['cumulative_percentage_customers'] = np.arange(1, len(df_true_sorted) + 1) / len(df_true_sorted)
    df_true_sorted['cumulative_percentage_ltv'] = df_true_sorted['true'].cumsum() / df_true_sorted['true'].sum()
    area_true = np.trapz(df_true_sorted['cumulative_percentage_ltv'],
                         df_true_sorted['cumulative_percentage_customers'])

    # 【3】绘图
    plt.figure(figsize=(10, 6))
    plt.plot(df_pred_sorted['cumulative_percentage_customers'],
             df_pred_sorted['cumulative_percentage_ltv'],
             label="Gain Curve (Predicted)", linewidth=2)
    plt.plot(df_true_sorted['cumulative_percentage_customers'],
             df_true_sorted['cumulative_percentage_ltv'],
             label="Ideal Gain Curve (Ground Truth Sorted)",
             linestyle='--', color='black', linewidth=2)
    plt.plot([0, 1], [0, 1], linestyle=':', color='gray', label="Random Model")

    plt.xlabel('Cumulative Percentage of Customers')
    plt.ylabel('Cumulative Percentage of Total LTV')
    plt.title(f'{head_name} Gain Chart')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.show()

    return area_pred

import sys
import os
import json
import tensorflow as tf
import tqdm

def create_hesitate_tf_dataset(dataset, MODEL_HOUR = 0):
    sample_batch = next(iter(dataset))
    sample_data = {k: v for k, v in sample_batch.items() if k not in ['b2_sale_amt_7d', 'total_pay_amount1']}

    def generator():
        for batch in dataset:
            hour = tf.cast(tf.gather(batch['user_sparse_features'],  indices=0, axis = 1) - 1, tf.int64)    # shape: (batch_size,)
            b2_7d = tf.cast(tf.reshape(batch.pop('b2_sale_amt_7d'), (-1, 1)), tf.float32)
            # 将 b2_7d 中小于 0 的值替换为 0
            b2_7d = tf.maximum(b2_7d, 0.0)
            
            total_amt_1h = tf.reshape(batch.pop('total_pay_amount1'), (-1, 1))

            # 只保留 hour 为 MODEL_HOUR 的记录
            hour_mask = tf.equal(hour, MODEL_HOUR)  # shape: (batch_size,)
            hour_mask = tf.reshape(hour_mask, (-1, 1))  # 广播成 (batch_size, 1)

            # 增加 犹豫人群 过滤mask 
            install_order_diff = tf.reshape(tf.cast(batch['all_install_order_7d_diff'], tf.int64), (-1, 1))
            
            # 构建 install_order_diff ∈ [0, 60] 的布尔 mask
            install_mask = tf.logical_or(
                install_order_diff < 0,
                install_order_diff > 60 * (MODEL_HOUR + 1)
            )
            combined_mask = tf.logical_and(hour_mask, install_mask)
            

            #  使用 hour_mask 筛选 batch 中的 对应小时窗口 
            selected_indices = tf.where(combined_mask)[:, 0]  # 获取 hour == 1 的样本索引
            batch = {k: tf.gather(v, selected_indices, axis=0) for k, v in batch.items()}  # 筛选 batch 中的样本
            b2_7d = tf.gather(b2_7d, selected_indices, axis=0)  # 保留 hour == 1 对应的标签
            total_amt_1h = tf.gather(total_amt_1h, selected_indices, axis=0)  # 保留 hour == 1 对应的标签
            
            # 将保留的样本和标签一起返回
            y_true_packed = tf.concat([b2_7d, total_amt_1h], axis=1)

            # y_true_packed = b2_7d
            yield batch, y_true_packed
        

    # 正确写法：output_signature 中保留每个字段的真实 shape
    output_signature = (
        {
            name: tf.TensorSpec(shape=(None,) + v.shape[1:], dtype=v.dtype)
            for name, v in sample_data.items()
        },
        tf.TensorSpec(shape=(None, 2), dtype=tf.float32)
    )

    return tf.data.Dataset.from_generator(generator, output_signature=output_signature)
